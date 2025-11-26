"""
Excel Dashboard Generator for Dehumidifier Dataset
Creates an Excel file with auto-updating graphs from the CSV dataset
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference, Series
from openpyxl.chart.axis import DateAxis
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# Configuration
DATA_DIR = r"C:\Users\User\OneDrive - itum.mrt.ac.lk\Desktop\Lecture Modules\Semester-4\Project 2\Project-Dehumidifier\Analyze Data"
CSV_FILENAME = "dehumidifier_data.csv"
EXCEL_FILENAME = "dehumidifier_dashboard.xlsx"

def create_dashboard(csv_path, excel_path):
    """Create Excel dashboard with graphs from CSV data"""
    
    # Load CSV data
    print(f"Loading data from: {csv_path}")
    df = pd.read_csv(csv_path)
    
    if len(df) == 0:
        print("Error: No data in CSV file!")
        return False
    
    print(f"Loaded {len(df)} records")
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # === SHEET 1: Raw Data ===
    ws_data = wb.create_sheet("Raw Data", 0)
    
    # Write dataframe to sheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)
    
    # Format header
    for cell in ws_data[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Set column widths
    ws_data.column_dimensions['A'].width = 22  # Timestamp
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws_data.column_dimensions[col].width = 12
    
    print("✓ Raw data imported")
    
    # === SHEET 2: Calculations ===
    ws_calc = wb.create_sheet("Calculations", 1)
    
    # Title styling
    ws_calc['A1'] = "System Performance Metrics"
    ws_calc['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_calc['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws_calc.merge_cells('A1:D1')
    ws_calc.row_dimensions[1].height = 25
    
    # Dehumidification Rate
    ws_calc['A3'] = "Dehumidification Rate Analysis"
    ws_calc['A3'].font = Font(bold=True, size=14)
    ws_calc['A3'].fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    ws_calc.merge_cells('A3:D3')
    
    ws_calc['A5'] = "Metric"
    ws_calc['B5'] = "Container 1"
    ws_calc['C5'] = "Container 2"
    ws_calc['D5'] = "Unit"
    for cell in ['A5', 'B5', 'C5', 'D5']:
        ws_calc[cell].font = Font(bold=True)
        ws_calc[cell].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    total_rows = len(df) + 1  # +1 for header
    
    # Average dehumidification rate (when fans are running)
    ws_calc['A6'] = "Avg Dehumid Rate (when ON)"
    ws_calc['B6'] = f'=AVERAGEIF(\'Raw Data\'!$G$2:$G${total_rows},1,\'Raw Data\'!$C$2:$C${total_rows})'
    ws_calc['C6'] = f'=AVERAGEIF(\'Raw Data\'!$G$2:$G${total_rows},1,\'Raw Data\'!$E$2:$E${total_rows})'
    ws_calc['D6'] = "% RH"
    
    # Max humidity recorded
    ws_calc['A7'] = "Max Humidity Recorded"
    ws_calc['B7'] = f'=MAX(\'Raw Data\'!$C$2:$C${total_rows})'
    ws_calc['C7'] = f'=MAX(\'Raw Data\'!$E$2:$E${total_rows})'
    ws_calc['D7'] = "% RH"
    
    # Min humidity recorded
    ws_calc['A8'] = "Min Humidity Recorded"
    ws_calc['B8'] = f'=MIN(\'Raw Data\'!$C$2:$C${total_rows})'
    ws_calc['C8'] = f'=MIN(\'Raw Data\'!$E$2:$E${total_rows})'
    ws_calc['D8'] = "% RH"
    
    # System Duty Cycle
    ws_calc['A10'] = "System Duty Cycle Analysis"
    ws_calc['A10'].font = Font(bold=True, size=14)
    ws_calc['A10'].fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    ws_calc.merge_cells('A10:D10')
    
    ws_calc['A12'] = "Metric"
    ws_calc['B12'] = "Value"
    ws_calc['C12'] = "Unit"
    for cell in ['A12', 'B12', 'C12']:
        ws_calc[cell].font = Font(bold=True)
        ws_calc[cell].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    ws_calc['A13'] = "Total Data Points"
    ws_calc['B13'] = f'=COUNTA(\'Raw Data\'!$G$2:$G${total_rows})'
    ws_calc['C13'] = "samples"
    
    ws_calc['A14'] = "Fans ON Count"
    ws_calc['B14'] = f'=COUNTIF(\'Raw Data\'!$G$2:$G${total_rows},1)'
    ws_calc['C14'] = "samples"
    
    ws_calc['A15'] = "Duty Cycle"
    ws_calc['B15'] = '=IF(B13>0,(B14/B13)*100,0)'
    ws_calc['B15'].number_format = '0.00'
    ws_calc['C15'] = "%"
    
    ws_calc['A16'] = "Total Running Time"
    ws_calc['B16'] = '=(B14*0.1)/60'
    ws_calc['B16'].number_format = '0.00'
    ws_calc['C16'] = "minutes"
    
    # Temperature Impact
    ws_calc['A18'] = "Temperature Impact Analysis"
    ws_calc['A18'].font = Font(bold=True, size=14)
    ws_calc['A18'].fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    ws_calc.merge_cells('A18:D18')
    
    ws_calc['A20'] = "Metric"
    ws_calc['B20'] = "Container 1"
    ws_calc['C20'] = "Container 2"
    ws_calc['D20'] = "Unit"
    for cell in ['A20', 'B20', 'C20', 'D20']:
        ws_calc[cell].font = Font(bold=True)
        ws_calc[cell].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    ws_calc['A21'] = "Avg Temp (Fans OFF)"
    ws_calc['B21'] = f'=AVERAGEIF(\'Raw Data\'!$G$2:$G${total_rows},0,\'Raw Data\'!$D$2:$D${total_rows})'
    ws_calc['C21'] = f'=AVERAGEIF(\'Raw Data\'!$G$2:$G${total_rows},0,\'Raw Data\'!$F$2:$F${total_rows})'
    ws_calc['D21'] = "°C"
    
    ws_calc['A22'] = "Avg Temp (Fans ON)"
    ws_calc['B22'] = f'=AVERAGEIF(\'Raw Data\'!$G$2:$G${total_rows},1,\'Raw Data\'!$D$2:$D${total_rows})'
    ws_calc['C22'] = f'=AVERAGEIF(\'Raw Data\'!$G$2:$G${total_rows},1,\'Raw Data\'!$F$2:$F${total_rows})'
    ws_calc['D22'] = "°C"
    
    ws_calc['A23'] = "Temperature Drop"
    ws_calc['B23'] = '=B21-B22'
    ws_calc['B23'].number_format = '0.00'
    ws_calc['C23'] = '=C21-C22'
    ws_calc['C23'].number_format = '0.00'
    ws_calc['D23'] = "°C"
    
    # Set column widths
    ws_calc.column_dimensions['A'].width = 30
    ws_calc.column_dimensions['B'].width = 15
    ws_calc.column_dimensions['C'].width = 15
    ws_calc.column_dimensions['D'].width = 10
    
    print("✓ Calculations sheet created")
    
    # === SHEET 3: Humidity Trends ===
    ws_humidity = wb.create_sheet("Humidity Trends", 2)
    
    ws_humidity['A1'] = "Humidity Trends Over Time"
    ws_humidity['A1'].font = Font(bold=True, size=14)
    
    # Create line chart for humidity
    chart_humidity = LineChart()
    chart_humidity.title = "Humidity Trends"
    chart_humidity.style = 13
    chart_humidity.y_axis.title = "Humidity (% RH)"
    chart_humidity.x_axis.title = "Sample Number"
    chart_humidity.height = 15
    chart_humidity.width = 25
    
    # Use limited data for chart (last 500 points or all if less)
    max_chart_rows = min(500, total_rows - 1)
    start_row = max(2, total_rows - max_chart_rows)
    
    # Humidity data (columns C and E from Raw Data)
    data_h1 = Reference(ws_data, min_col=3, min_row=start_row, max_row=total_rows)
    data_h2 = Reference(ws_data, min_col=5, min_row=start_row, max_row=total_rows)
    
    chart_humidity.add_data(data_h1, titles_from_data=False)
    chart_humidity.add_data(data_h2, titles_from_data=False)
    
    chart_humidity.series[0].graphicalProperties.line.solidFill = "0070C0"
    chart_humidity.series[0].graphicalProperties.line.width = 20000
    chart_humidity.series[1].graphicalProperties.line.solidFill = "FF6B35"
    chart_humidity.series[1].graphicalProperties.line.width = 20000
    
    chart_humidity.legend.position = 'r'
    
    ws_humidity.add_chart(chart_humidity, "A3")
    
    print("✓ Humidity trends chart created")
    
    # === SHEET 4: Temperature Impact ===
    ws_temp = wb.create_sheet("Temperature Impact", 3)
    
    ws_temp['A1'] = "Temperature During Dehumidification"
    ws_temp['A1'].font = Font(bold=True, size=14)
    
    # Create line chart for temperature
    chart_temp = LineChart()
    chart_temp.title = "Temperature Trends"
    chart_temp.style = 12
    chart_temp.y_axis.title = "Temperature (°C)"
    chart_temp.x_axis.title = "Sample Number"
    chart_temp.height = 15
    chart_temp.width = 25
    
    # Temperature data (columns D and F from Raw Data)
    data_t1 = Reference(ws_data, min_col=4, min_row=start_row, max_row=total_rows)
    data_t2 = Reference(ws_data, min_col=6, min_row=start_row, max_row=total_rows)
    
    chart_temp.add_data(data_t1, titles_from_data=False)
    chart_temp.add_data(data_t2, titles_from_data=False)
    
    chart_temp.series[0].graphicalProperties.line.solidFill = "C00000"
    chart_temp.series[0].graphicalProperties.line.width = 20000
    chart_temp.series[1].graphicalProperties.line.solidFill = "FF9900"
    chart_temp.series[1].graphicalProperties.line.width = 20000
    
    chart_temp.legend.position = 'r'
    
    ws_temp.add_chart(chart_temp, "A3")
    
    # Add summary metrics
    ws_temp['A25'] = "Temperature Impact Summary"
    ws_temp['A25'].font = Font(bold=True, size=12)
    
    ws_temp['A27'] = "Container 1 Temp Drop:"
    ws_temp['B27'] = "=Calculations!B23"
    ws_temp['B27'].number_format = '0.00" °C"'
    
    ws_temp['A28'] = "Container 2 Temp Drop:"
    ws_temp['B28'] = "=Calculations!C23"
    ws_temp['B28'].number_format = '0.00" °C"'
    
    print("✓ Temperature impact chart created")
    
    # === SHEET 5: System Activity ===
    ws_activity = wb.create_sheet("System Activity", 4)
    
    ws_activity['A1'] = "System Activity Dashboard"
    ws_activity['A1'].font = Font(bold=True, size=14)
    
    # Create line chart for system status
    chart_activity = LineChart()
    chart_activity.title = "Fans & Peltier Activity"
    chart_activity.style = 11
    chart_activity.y_axis.title = "Status (0=OFF, 1=ON)"
    chart_activity.x_axis.title = "Sample Number"
    chart_activity.height = 12
    chart_activity.width = 25
    
    # System activity data (columns G and H from Raw Data)
    data_fans = Reference(ws_data, min_col=7, min_row=start_row, max_row=total_rows)
    data_peltier = Reference(ws_data, min_col=8, min_row=start_row, max_row=total_rows)
    
    chart_activity.add_data(data_fans, titles_from_data=False)
    chart_activity.add_data(data_peltier, titles_from_data=False)
    
    chart_activity.series[0].graphicalProperties.line.solidFill = "00B050"
    chart_activity.series[0].graphicalProperties.line.width = 25000
    chart_activity.series[1].graphicalProperties.line.solidFill = "7030A0"
    chart_activity.series[1].graphicalProperties.line.width = 25000
    
    chart_activity.legend.position = 'r'
    
    ws_activity.add_chart(chart_activity, "A3")
    
    # Add duty cycle summary
    ws_activity['A20'] = "System Performance Summary"
    ws_activity['A20'].font = Font(bold=True, size=12)
    ws_activity['A20'].fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    
    ws_activity['A22'] = "Duty Cycle:"
    ws_activity['B22'] = "=Calculations!B15"
    ws_activity['B22'].number_format = '0.00"%"'
    ws_activity['B22'].font = Font(size=14, bold=True, color="0070C0")
    
    ws_activity['A23'] = "Total Running Time:"
    ws_activity['B23'] = "=Calculations!B16"
    ws_activity['B23'].number_format = '0.00" min"'
    ws_activity['B23'].font = Font(size=14, bold=True)
    
    ws_activity['A24'] = "Total Data Points:"
    ws_activity['B24'] = "=Calculations!B13"
    ws_activity['B24'].number_format = '#,##0'
    
    ws_activity['A25'] = "Data Collection Time:"
    ws_activity['B25'] = '=(Calculations!B13*0.1)/60'
    ws_activity['B25'].number_format = '0.00" min"'
    
    print("✓ System activity chart created")
    
    # Save workbook
    wb.save(excel_path)
    print(f"\n✓ Dashboard saved: {excel_path}")
    
    return True

def main():
    csv_path = os.path.join(DATA_DIR, CSV_FILENAME)
    excel_path = os.path.join(DATA_DIR, EXCEL_FILENAME)
    
    # Check if CSV exists
    if not os.path.exists(csv_path):
        print(f"Error: CSV file not found!")
        print(f"Expected location: {csv_path}")
        print("\nPlease run data_logger.py first to collect data.")
        return
    
    print("="*60)
    print("Creating Excel Dashboard with Graphs")
    print("="*60)
    print()
    
    success = create_dashboard(csv_path, excel_path)
    
    if success:
        print("\n" + "="*60)
        print("✓ SUCCESS! Dashboard created with:")
        print("  - Humidity Trends Chart")
        print("  - Temperature Impact Chart")
        print("  - System Activity Chart")
        print("  - Dehumidification Rate Analysis")
        print("  - System Duty Cycle Analysis")
        print("="*60)
        print(f"\nOpen this file: {excel_path}")
        print("\nNote: Run this script again after collecting more data")
        print("to update the dashboard with latest information.")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\nError: {e}")
        print("\nMake sure you have installed:")
        print("  pip install pandas openpyxl")
